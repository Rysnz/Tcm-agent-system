import os
import uuid
import traceback
from typing import Dict, Any, List
from django.core.files.uploadedfile import UploadedFile
from apps.knowledge.models import KnowledgeBase, Document, Paragraph
from apps.knowledge.vector.pg_vector import PGVectorStore
import logging

logger = logging.getLogger('apps')

class DocumentProcessor:
    
    def __init__(self, knowledge_base_id: str):
        self.knowledge_base_id = knowledge_base_id
        # 获取知识库配置的向量模型，默认为shibing624/text2vec-base-chinese
        try:
            kb = KnowledgeBase.objects.get(id=knowledge_base_id)
            if hasattr(kb, 'embedding_model') and kb.embedding_model:
                self.embedding_model = kb.embedding_model
                self.dimension = getattr(kb, 'embedding_dimension', 768)  # shibing624/text2vec-base-chinese的维度是768
            else:
                self.embedding_model = 'shibing624/text2vec-base-chinese'
                self.dimension = 768
        except KnowledgeBase.DoesNotExist:
            self.embedding_model = 'shibing624/text2vec-base-chinese'
            self.dimension = 768
        
        self.vector_store = PGVectorStore({
            'knowledge_base_id': knowledge_base_id,
            'embedding_model': self.embedding_model,
            'dimension': self.dimension
        })
        
        # 确保向量索引存在
        self.vector_store.ensure_index_exists()
    
    def process_file(self, file: UploadedFile) -> Document:
        """
        处理上传的文件，包括提取文本、保存段落和生成向量
        :param file: 上传的文件
        :return: 处理后的文档对象
        """
        file_name = file.name
        file_size = file.size
        file_type = self._get_file_type(file_name)
        
        file_path = self._save_file(file)
        
        knowledge_base = KnowledgeBase.objects.get(id=self.knowledge_base_id)
        
        document = Document.objects.create(
            knowledge_base=knowledge_base,
            name=file_name,
            file_type=file_type,
            file_size=file_size,
            file_path=file_path,
            status='processing'
        )
        
        try:
            # 提取文本
            paragraphs = self._extract_text(file_path, file_type)
            
            # 先保存文本处理结果
            document.char_count = sum(len(p['content']) for p in paragraphs)
            document.paragraph_count = len(paragraphs)
            document.progress = 50
            document.save()
            
            try:
                # 保存段落和生成向量
                self._process_paragraphs(document, paragraphs)
                document.status = 'completed'
                document.progress = 100
                document.save()
            except Exception as vector_error:
                # 向量化失败时，文档状态设置为partially_completed
                document.status = 'partially_completed'
                document.save()
                logger.error(f"向量化处理失败: {document.id}, 错误类型: {type(vector_error).__name__}, 错误信息: {str(vector_error)}")
            
        except Exception as e:
            document.status = 'failed'
            document.progress = 100
            document.save()
            logger.error(f"文档处理失败: {document.id}, 错误类型: {type(e).__name__}, 错误信息: {str(e)}, 堆栈: {traceback.format_exc()}")
            raise e
        
        return document
    
    def _get_file_type(self, file_name: str) -> str:
        """
        获取文件类型
        :param file_name: 文件名
        :return: 文件类型
        """
        ext = os.path.splitext(file_name)[1].lower()
        type_map = {
            '.pdf': 'pdf',
            '.doc': 'doc',
            '.docx': 'docx',
            '.txt': 'txt',
            '.md': 'md',
            '.xlsx': 'xlsx',
            '.xls': 'xls'
        }
        return type_map.get(ext, 'unknown')
    
    def _save_file(self, file: UploadedFile) -> str:
        """
        保存上传的文件
        :param file: 上传的文件
        :return: 文件路径
        """
        upload_dir = f'media/documents/{self.knowledge_base_id}'
        os.makedirs(upload_dir, exist_ok=True)
        
        file_path = os.path.join(upload_dir, f"{uuid.uuid4()}_{file.name}")
        with open(file_path, 'wb+') as destination:
            for chunk in file.chunks():
                destination.write(chunk)
        
        logger.info(f"文件保存成功: {file_path}, 大小: {file.size}字节")
        return file_path
    
    def _extract_text(self, file_path: str, file_type: str) -> list:
        """
        提取文件文本
        :param file_path: 文件路径
        :param file_type: 文件类型
        :return: 段落列表
        """
        logger.info(f"开始提取文件文本: {file_path}, 类型: {file_type}")
        paragraphs = []
        
        try:
            if file_type == 'pdf':
                paragraphs = self._extract_pdf(file_path)
            elif file_type in ['doc', 'docx']:
                paragraphs = self._extract_word(file_path)
            elif file_type in ['txt', 'md']:
                paragraphs = self._extract_text_file(file_path)
            elif file_type in ['xlsx', 'xls']:
                paragraphs = self._extract_excel(file_path)
            else:
                logger.warning(f"不支持的文件类型: {file_type}")
                
            logger.info(f"文件文本提取完成: {file_path}, 提取段落数: {len(paragraphs)}")
            return paragraphs
        except Exception as e:
            logger.error(f"文件文本提取失败: {file_path}, 错误: {str(e)}")
            raise e
    
    def _extract_pdf(self, file_path: str) -> list:
        """
        提取PDF文件文本
        :param file_path: PDF文件路径
        :return: 段落列表
        """
        from pypdf import PdfReader
        paragraphs = []
        
        with open(file_path, 'rb') as file:
            reader = PdfReader(file)
            total_pages = len(reader.pages)
            logger.info(f"PDF文件总页数: {total_pages}")
            
            for page_num, page in enumerate(reader.pages):
                text = page.extract_text()
                if text.strip():
                    # 使用智能分段算法处理每页文本
                    page_paragraphs = self._smart_split(text, max_length=1000, title=f'第{page_num + 1}页')
                    for para in page_paragraphs:
                        para['page_number'] = page_num + 1
                        para['meta'] = {'source': 'pdf', 'page': page_num + 1, 'total_pages': total_pages}
                    paragraphs.extend(page_paragraphs)
        
        return paragraphs
    
    def _extract_word(self, file_path: str) -> list:
        """
        提取Word文件文本
        :param file_path: Word文件路径
        :return: 段落列表
        """
        paragraphs = []
        
        # 检查文件后缀，分别处理.doc和.docx
        file_ext = os.path.splitext(file_path)[1].lower()
        
        if file_ext == '.docx':
            # 处理.docx文件，使用python-docx库
            from docx import Document
            doc = Document(file_path)
            
            # 合并所有段落文本，然后使用智能分段
            full_text = '\n'.join([para.text for para in doc.paragraphs if para.text.strip()])
            if full_text.strip():
                paragraphs = self._smart_split(full_text, max_length=1000, title=os.path.basename(file_path))
                for para in paragraphs:
                    para['meta'] = {'source': 'word'}
        elif file_ext == '.doc':
            # 处理.doc文件，添加详细的错误日志
            logger.info(f"开始处理.doc文件: {file_path}")
            
            # 尝试多种方法处理.doc文件
            methods = []
            
            # 方法1: 使用win32com.client（Windows环境下可用）
            def try_win32com():
                logger.info(f"尝试使用win32com处理.doc文件: {file_path}")
                try:
                    import win32com.client
                    import pythoncom
                    
                    # 初始化COM
                    pythoncom.CoInitialize()
                    
                    # 创建Word应用对象
                    word = win32com.client.Dispatch('Word.Application')
                    word.Visible = False
                    
                    # 打开文档，使用绝对路径
                    abs_file_path = os.path.abspath(file_path)
                    doc = word.Documents.Open(abs_file_path)
                    
                    # 提取文本
                    text = doc.Content.Text
                    
                    # 关闭文档
                    doc.Close()
                    word.Quit()
                    
                    logger.info(f"使用win32com成功提取.doc文件内容: {file_path}")
                    return text
                except Exception as win32_error:
                    logger.error(f"使用win32com处理.doc文件失败: {file_path}, 错误: {str(win32_error)}")
                    return None
            methods.append(try_win32com)
            
            # 方法2: 尝试使用pymupdf（作为备选方案）
            def try_pymupdf():
                logger.info(f"尝试使用pymupdf处理.doc文件: {file_path}")
                try:
                    import fitz
                    # 注意：pymupdf主要用于PDF，但也可以尝试读取某些.doc文件
                    doc = fitz.open(file_path)
                    full_text = ''
                    for page_num in range(len(doc)):
                        page = doc.load_page(page_num)
                        text = page.get_text()
                        if text.strip():
                            full_text += text + '\n'
                    doc.close()
                    if full_text.strip():
                        logger.info(f"使用pymupdf成功提取.doc文件内容: {file_path}")
                        return full_text
                    else:
                        logger.warning(f"使用pymupdf提取.doc文件内容为空: {file_path}")
                        return None
                except Exception as e:
                    logger.error(f"使用pymupdf处理.doc文件失败: {file_path}, 错误: {str(e)}")
                    return None
            methods.append(try_pymupdf)
            
            # 方法3: 直接报错，告知用户.doc文件处理需要额外的库
            def fallback_error():
                error_msg = f"所有方法都无法处理.doc文件: {file_path}, 请尝试将文件转换为.docx格式后再上传"
                logger.error(error_msg)
                raise Exception(error_msg)
            methods.append(fallback_error)
            
            # 尝试所有方法，直到找到可用的
            text = None
            for method in methods:
                text = method()
                if text is not None:
                    break
            
            if text and text.strip():
                paragraphs = self._smart_split(text, max_length=1000, title=os.path.basename(file_path))
                for para in paragraphs:
                    para['meta'] = {'source': 'word'}
        
        return paragraphs
    
    def _extract_text_file(self, file_path: str) -> list:
        """
        提取文本文件文本
        :param file_path: 文本文件路径
        :return: 段落列表
        """
        paragraphs = []
        
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                content = file.read()
                if content.strip():
                    # 使用智能分段算法处理文本文件，避免过多短段落
                    paragraphs = self._smart_split(content, max_length=1000, title=os.path.basename(file_path))
                    for para in paragraphs:
                        para['meta'] = {'source': 'text'}
        except UnicodeDecodeError:
            # 尝试使用其他编码
            logger.warning(f"UTF-8编码读取失败，尝试使用GBK编码: {file_path}")
            with open(file_path, 'r', encoding='gbk') as file:
                content = file.read()
                if content.strip():
                    paragraphs = self._smart_split(content, max_length=1000, title=os.path.basename(file_path))
                    for para in paragraphs:
                        para['meta'] = {'source': 'text'}
        
        return paragraphs
    
    def _extract_excel(self, file_path: str) -> list:
        """
        提取Excel文件文本
        :param file_path: Excel文件路径
        :return: 段落列表
        """
        import openpyxl
        paragraphs = []
        
        try:
            wb = openpyxl.load_workbook(file_path)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                for row_num, row in enumerate(sheet.iter_rows(values_only=True)):
                    row_text = ' '.join([str(cell) if cell else '' for cell in row])
                    if row_text.strip():
                        paragraphs.append({
                            'content': row_text,
                            'title': f'{sheet_name}第{row_num + 1}行',
                            'meta': {'source': 'excel', 'sheet': sheet_name, 'row': row_num + 1}
                        })
        except Exception as e:
            logger.error(f"Excel文件处理失败: {file_path}, 错误: {str(e)}")
            raise e
        
        return paragraphs
    
    def _smart_split(self, text: str, max_length: int = 1000, title: str = '') -> list:
        """
        智能分段算法
        :param text: 待分段的文本
        :param max_length: 最大段落长度
        :param title: 文档标题
        :return: 分段后的段落列表
        """
        import re
        
        paragraphs = []
        text = text.strip()
        
        if not text:
            return paragraphs
        
        # 优先按标点符号分段，确保语义完整
        # 匹配规则：1到max_length个字符，后跟中文标点、英文标点或换行符
        pattern = rf'.{{1,{max_length}}}[。！？!?;；\\n]'
        
        matches = re.finditer(pattern, text)
        
        start = 0
        for match in matches:
            end = match.end()
            content = text[start:end].strip()
            if content:
                paragraphs.append({
                    'content': content,
                    'title': title
                })
            start = end
        
        # 处理剩余文本
        if start < len(text):
            remaining = text[start:].strip()
            if remaining:
                # 如果剩余文本过长，再次分段
                if len(remaining) > max_length:
                    # 递归处理剩余文本
                    remaining_paragraphs = self._smart_split(remaining, max_length, title)
                    paragraphs.extend(remaining_paragraphs)
                else:
                    paragraphs.append({
                        'content': remaining,
                        'title': title
                    })
        
        return paragraphs
    
    def _process_paragraphs(self, document: Document, paragraphs: List[Dict[str, Any]]):
        """
        处理段落，包括保存和向量化
        :param document: 文档对象
        :param paragraphs: 段落列表
        :return: 处理后的段落列表
        """
        # 准备数据
        texts = [para['content'] for para in paragraphs]
        metadatas = []
        
        for para in paragraphs:
            metadata = para.copy()
            metadata.pop('content', None)  # 移除content字段，因为它已经作为texts传递
            metadata['document_id'] = str(document.id)
            metadata['knowledge_id'] = str(document.knowledge_base.id)
            metadata['source_type'] = 'paragraph'
            metadata['is_active'] = True
            metadatas.append(metadata)
        
        # 使用向量存储的batch_save方法处理
        # 先保存段落和向量
        self.vector_store.batch_save([
            {
                'text': text,
                **meta
            } for text, meta in zip(texts, metadatas)
        ])
        
        # 更新文档进度
        document.progress = 80
        document.save()
        
        logger.info(f"段落处理完成: 文档ID {document.id}, 处理段落数: {len(paragraphs)}")
